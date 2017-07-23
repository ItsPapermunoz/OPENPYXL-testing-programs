from openpyxl import Workbook
import os


wb = Workbook()
ws = wb.active
ws1 = wb.create_sheet("Sheet 2")
c = ws['A4']
a = ws['B4']
c.value = "Hello this is a string test..."
a.value = 50
wb.save("Testing.xlsx")
