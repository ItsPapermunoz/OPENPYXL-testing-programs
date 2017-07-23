from openpyxl import load_workbook
from openpyxl import Workbook
wb = load_workbook("testing.xlsx")
ws = wb.active
cella = ws["A4"]
print(cella.value)
